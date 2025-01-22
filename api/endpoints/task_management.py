from datetime import date, datetime
from fastapi import APIRouter, Depends, Form, Query, status, HTTPException, Response, UploadFile
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import insert
from database import get_db
from models.response import GeneralDataPaginateResponse, GeneralDataResponse
# from services.user.user_group_service import UserGroupService
from models.task_management import TaskManagement
from repositories.task_management_repository import TaskManagementRepository
from services.task_management_service import TaskManagementRepository, TaskManagementService
from services.user_service import UserService
from utils.authentication import Authentication
from utils.manual import get_total_pages
import os
from typing import Optional, List
# from models.project import Project
# from models.user.user_group import UserGroup
from io import BytesIO
from openpyxl.utils import get_column_letter
# from openpyxl import load_workbook
# from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
import math
router = APIRouter()

@router.post("", response_model=GeneralDataResponse, status_code=status.HTTP_201_CREATED)
def create_task_management(
    task_management: str = Form(..., min_length=1, max_length=256),
    description: Optional[str] = Form(None, min_length=0, max_length=512),
    date: str = Form(...),
    db: Session = Depends(get_db),
    payload=Depends(Authentication())
):
    """
    Membuat Tugas harian
    """

    user_id_active = payload.get("uid", None)

    user_service = UserService(db)
    task_management_service = TaskManagementService(db)  # Buat instance dengan db
    user_active = user_service.user_repository.read_user(user_id_active)

    try:
        # Buat instance TaskManagement
        task_management_model = TaskManagement(
            user_id=user_active.id,  # Pastikan ini adalah ID, bukan objek
            task_management=task_management,
            description=description,
            date=date,
            priority=False,
        )
        # Panggil method pada instance task_management_service
        task_management = task_management_service.create_task_management(task_management=task_management_model)
    except ValueError as error:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(error))

    result = {
        'id': task_management.id,
    }
    data_response = GeneralDataResponse(
        code=status.HTTP_201_CREATED,
        status="OK",
        data=result,
    )
    return data_response

@router.get("", response_model=GeneralDataPaginateResponse, status_code=status.HTTP_200_OK)
def read_tasks(
    date: Optional[date] = Query(None),
    sort_by: Optional[str] = Query(None),
    sort_order: Optional[str] = Query(None),
    task_management_id: Optional[str] = Query(None),
    user_id: Optional[str] = Query(None),
    offset: Optional[int] = Query(1, ge=1), 
    size: Optional[int] = Query(10, ge=1),  # Default size set to 10
    db: Session = Depends(get_db), 
    payload = Depends(Authentication())
):
    """
    Read All Tasks or Task by Date

    - need login
    """
    user_id_active = payload.get("uid", None)
    user_role = payload.get("role", None)

    # Services
    task_management_service = TaskManagementService(db)
    user_service = UserService(db)

    # Validasi user aktif
    user_active = user_service.user_repository.read_user(user_id_active)
    
    # Fetch tasks based on the filters and date, but only tasks owned by the current user
    task_managements = task_management_service.task_management_repository.read_task_managements(
        sort_by=sort_by,
        sort_order=sort_order,
        offset=offset,
        size=size,
        user_id=user_id_active,
        task_management_id=task_management_id
    )  

    if not task_managements:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Data not found")

    # Prepare data for the response
    datas = []
    for task in task_managements:
        datas.append({
            'id': task.id,  # Correct the task reference
            'user': task.user.name,  # Accessing the correct task object
            'task_management': task.task_management,
            'description': task.description,            
            'priority': task.priority,
            'date': task.date.strftime('%Y-%m-%d') if task.date else None,            
            'created_at': str(task.created_at),
            'updated_at': str(task.updated_at),
        })

    # Prepare metadata
    count = task_management_service.task_management_repository.count_task_managements(
        user_id=user_id_active,task_management_id=task_management_id  
    )

    total_pages = get_total_pages(size, count)
    meta = {
        "size": size,
        "total": count,
        "total_pages": total_pages,
        "offset": offset,
    }
    
    # Prepare response
    status_code = status.HTTP_200_OK
    data_response = GeneralDataPaginateResponse(
        code=status_code,
        status="OK",
        data=datas,
        meta=meta,
    )
    return data_response  # No need for JSONResponse here, FastAPI handles it automatically


@router.get("/{task_management_id}", response_model=GeneralDataResponse, status_code=status.HTTP_200_OK)
def read_task_management(
    task_management_id: str,
    db: Session = Depends(get_db), 
    payload: dict = Depends(Authentication())
):
    """
    Read Task Management Detail.

    - Authentication required.
    """
    # Ambil user ID aktif dari payload
    user_id_active = payload.get("uid")
    if not user_id_active:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="User not authenticated")

    # Inisialisasi service
    user_service = UserService(db)
    task_management_service = TaskManagementService(db)

    # Validasi user aktif
    user_active = user_service.user_repository.read_user(user_id_active)
    if not user_active:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")


    # Ambil data task management berdasarkan ID
    task_management = task_management_service.task_management_repository.read_task_management(task_management_id)
    if not task_management:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Task management data not found")
    
    # Validasi apakah task yang akan dihapus milik user yang sedang login
    if task_management.user_id != user_id_active:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Tidak dapat melihat task milik orang lain")

    # Bangun respons untuk user terkait task management
    user = {
        'id': task_management.user.id,
        'name': task_management.user.name,
    } if task_management.user else {}

    # Format respons umum
    data_response = GeneralDataResponse(
        code=status.HTTP_200_OK,
        status="OK",
        data={
            'id': task_management.id,
            'user': user,
            'task_management': task_management.task_management,
            'description': task_management.description,
            'date': task_management.date,
            'priority': task_management.priority,
            'created_at': task_management.created_at.isoformat(),
            'updated_at': task_management.updated_at.isoformat(),
        },
    )
    
    return data_response  

@router.patch("/{task_management_id}", response_model=GeneralDataResponse, status_code=status.HTTP_200_OK)
async def update_task(
    task_management_id: str,
    task_management: str = Form(..., min_length=1, max_length=256),
    description: str = Form(None, min_length=0, max_length=512),    
    date: date = Form(...),
    priority: bool = Form(...),
    db: Session = Depends(get_db), 
    payload = Depends(Authentication())
):
    """
    Update Task

    - should login
    """
    user_id_active = payload.get("uid", None)

    if not user_id_active:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="User not authenticated")

    user_service = UserService(db)
    task_management_service = TaskManagementService(db)

    user_active = user_service.user_repository.read_user(user_id_active)

    if not user_active:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    
    # Cari task berdasarkan ID
    exist_task = task_management_service.task_management_repository.read_task_management(task_management_id)
    if not exist_task:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Task not found")
    
    try:
        # Perbarui data task
        updated_task_management = TaskManagement(
            id=task_management_id,  # Pastikan ID yang sesuai digunakan
            task_management=task_management,
            description=description,
            date=date,
            priority=priority,
        )

        # Panggil update_task_management dengan data yang sudah diperbarui
        updated_task = task_management_service.task_management_repository.update_task_management(updated_task_management)
    except ValueError as error:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(error))

    status_code = status.HTTP_200_OK
    data = {
        'id': updated_task.id,
    }

    data_response = GeneralDataResponse(
        code=status_code,
        status="OK",
        data=data,
    )
    response = JSONResponse(content=data_response.model_dump(), status_code=status_code)
    return response


@router.delete("/{task_management_id}", response_model=GeneralDataResponse, status_code=status.HTTP_200_OK)
async def delete_task(
    task_management_id: str,
    db: Session = Depends(get_db), 
    payload = Depends(Authentication())
):
    """
        Delete Task
        
        - harus login
        - tidak bisa menghapus task milik orang lain
    """
    user_id_active = payload.get("uid", None)

    user_service = UserService(db)
    task_management_service = TaskManagementService(db)
    
    # Cari task yang akan dihapus
    exist_task = task_management_service.task_management_repository.read_task_management(task_management_id)
    if not exist_task:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Task tidak ditemukan")
    
    # Validasi apakah task yang akan dihapus milik user yang sedang login
    if exist_task.user_id != user_id_active:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Tidak dapat menghapus task milik orang lain")

    # Hapus task
    try:
        task_management_service.task_management_repository.delete_task_management(task_management_id)
        # Tidak perlu db.commit() jika commit sudah dilakukan dalam delete_task_management
    except Exception as error:
        db.rollback()  # Jika terjadi error, rollback perubahan
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(error))

    # Persiapkan respons
    status_code = status.HTTP_200_OK
    data = {
        'id': task_management_id,
    }

    data_response = GeneralDataResponse(
        code=status_code,
        status="OK",
        data=data,
    )

    return data_response  # FastAPI akan otomatis menangani serialisasi ke JSON
