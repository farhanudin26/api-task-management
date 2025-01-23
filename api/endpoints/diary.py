from datetime import date, datetime
from fastapi import APIRouter, Depends, Form, Query, status, HTTPException, Response, UploadFile
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import insert
from database import get_db
from models.diary import Diary
from models.response import GeneralDataPaginateResponse, GeneralDataResponse
# from services.user.user_group_service import UserGroupService
from services.diary_service import DiaryService
from services.user_service import UserService
from utils.authentication import Authentication
from utils.manual import get_total_pages
import os
from typing import Optional, List
# from models.project import Project
# from models.user.user_group import UserGroup
from io import BytesIO
from openpyxl.utils import get_column_letter
# from openpyxl import load_workbook
# from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
import math
router = APIRouter()

@router.post("", response_model=GeneralDataResponse, status_code=status.HTTP_201_CREATED)
def create_diary(
    diary: str = Form(..., min_length=1, max_length=512),
    date: str = Form(...),
    db: Session = Depends(get_db),
    payload=Depends(Authentication())
):
    """
    Membuat Catatan Harian
    """

    user_id_active = payload.get("uid", None)

    # Inisialisasi service
    user_service = UserService(db)
    diary_service = DiaryService(db)

    # Validasi user aktif
    user_active = user_service.user_repository.read_user(user_id_active)

    try:
        # Buat instance Diary (model ORM)
        diary_model = Diary(
            user_id=user_active.id,
            diary=diary,  # diary sebagai teks
            date=date,
        )
        
        # Gunakan service untuk menyimpan diary ke database
        diary = diary_service.create_diary(diary=diary_model)
    except ValueError as error:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(error))

    # Siapkan response
    result = {
        'id': diary.id,
    }
    data_response = GeneralDataResponse(
        code=status.HTTP_201_CREATED,
        status="OK",
        data=result,
    )
    return data_response


@router.get("", response_model=GeneralDataPaginateResponse, status_code=status.HTTP_200_OK)
def read_diarys(
    date: Optional[date] = Query(None),
    sort_by: Optional[str] = Query(None),
    sort_order: Optional[str] = Query(None),
    diary_id: Optional[str] = Query(None),
    offset: Optional[int] = Query(1, ge=1), 
    size: Optional[int] = Query(10, ge=1),  # Default size set to 10
    db: Session = Depends(get_db), 
    payload = Depends(Authentication())
):
    """
get catatan harian
    """
    user_id_active = payload.get("uid", None)
    user_role = payload.get("role", None)

    # Services
    diary_service = DiaryService(db)
    user_service = UserService(db)

    # Validasi user aktif
    user_active = user_service.user_repository.read_user(user_id_active)
    
    # Fetch diarys based on the filters and date, but only diarys owned by the current user
    diarys = diary_service.diary_repository.read_diarys(
        sort_by=sort_by,
        sort_order=sort_order,
        offset=offset,
        size=size,
        user_id=user_id_active,
        diary_id=diary_id
    )  

    if not diarys:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Data not found")

    # Prepare data for the response
    datas = []
    for diary in diarys:
        datas.append({
            'id': diary.id,  # Correct the diary reference
            'user': diary.user.name,  # Accessing the correct diary object
            'diary': diary.diary,
            'date': diary.date.strftime('%Y-%m-%d') if diary.date else None,            
            'created_at': str(diary.created_at),
            'updated_at': str(diary.updated_at),
        })

    # Prepare metadata
    count = diary_service.diary_repository.count_diarys(
        user_id=user_id_active,diary_id=diary_id  
    )

    total_pages = get_total_pages(size, count)
    meta = {
        "size": size,
        "total": count,
        "total_pages": total_pages,
        "offset": offset,
    }
    
    # Prepare response
    status_code = status.HTTP_200_OK
    data_response = GeneralDataPaginateResponse(
        code=status_code,
        status="OK",
        data=datas,
        meta=meta,
    )
    return data_response  # No need for JSONResponse here, FastAPI handles it automatically


@router.get("/{diary_id}", response_model=GeneralDataResponse, status_code=status.HTTP_200_OK)
def read_diary(
    diary_id: str,
    db: Session = Depends(get_db), 
    payload: dict = Depends(Authentication())
):
    """
    Read diary Management Detail.

    - Authentication required.
    """
    # Ambil user ID aktif dari payload
    user_id_active = payload.get("uid")
    if not user_id_active:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="User not authenticated")

    # Inisialisasi service
    user_service = UserService(db)
    diary_service = DiaryService(db)

    # Validasi user aktif
    user_active = user_service.user_repository.read_user(user_id_active)
    if not user_active:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")


    # Ambil data diary management berdasarkan ID
    diary = diary_service.diary_repository.read_diary(diary_id)
    if not diary:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="diary management data not found")
    
    # Validasi apakah diary yang akan dihapus milik user yang sedang login
    if diary.user_id != user_id_active:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Tidak dapat melihat diary milik orang lain")

    # Bangun respons untuk user terkait diary management
    user = {
        'id': diary.user.id,
        'name': diary.user.name,
    } if diary.user else {}

    # Format respons umum
    data_response = GeneralDataResponse(
        code=status.HTTP_200_OK,
        status="OK",
        data={
            'id': diary.id,
            'user': user,
            'diary': diary.diary,
            'date': diary.date,
            'created_at': diary.created_at.isoformat(),
            'updated_at': diary.updated_at.isoformat(),
        },
    )
    
    return data_response  

@router.patch("/{diary_id}", response_model=GeneralDataResponse, status_code=status.HTTP_200_OK)
async def update_diary(
    diary_id: str,
    diary: str = Form(..., min_length=1, max_length=256),
    date: date = Form(...),
    db: Session = Depends(get_db), 
    payload = Depends(Authentication())
):
    """
    Update diary

    - should login
    """
    user_id_active = payload.get("uid", None)

    if not user_id_active:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="User not authenticated")

    user_service = UserService(db)
    diary_service = DiaryService(db)

    user_active = user_service.user_repository.read_user(user_id_active)

    if not user_active:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    
    # Cari diary berdasarkan ID
    exist_diary = diary_service.diary_repository.read_diary(diary_id)
    if not exist_diary:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="diary not found")
    
    try:
        # Perbarui data diary
        updated_diary = Diary(
            id=diary_id,  # Pastikan ID yang sesuai digunakan
            diary=diary,
            date=date,
        )

        # Panggil update_diary dengan data yang sudah diperbarui
        updated_diary = diary_service.diary_repository.update_diary(updated_diary)
    except ValueError as error:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(error))

    status_code = status.HTTP_200_OK
    data = {
        'id': updated_diary.id,
    }

    data_response = GeneralDataResponse(
        code=status_code,
        status="OK",
        data=data,
    )
    response = JSONResponse(content=data_response.model_dump(), status_code=status_code)
    return response


@router.delete("/{diary_id}", response_model=GeneralDataResponse, status_code=status.HTTP_200_OK)
async def delete_diary(
    diary_id: str,
    db: Session = Depends(get_db), 
    payload = Depends(Authentication())
):
    """
        Delete diary
        
        - harus login
        - tidak bisa menghapus diary milik orang lain
    """
    user_id_active = payload.get("uid", None)

    user_service = UserService(db)
    diary_service = DiaryService(db)
    
    # Cari diary yang akan dihapus
    exist_diary = diary_service.diary_repository.read_diary(diary_id)
    if not exist_diary:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="diary tidak ditemukan")
    
    # Validasi apakah diary yang akan dihapus milik user yang sedang login
    if exist_diary.user_id != user_id_active:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Tidak dapat menghapus diary milik orang lain")

    # Hapus diary
    try:
        diary_service.diary_repository.delete_diary(diary_id)
        # Tidak perlu db.commit() jika commit sudah dilakukan dalam delete_diary
    except Exception as error:
        db.rollback()  # Jika terjadi error, rollback perubahan
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(error))

    # Persiapkan respons
    status_code = status.HTTP_200_OK
    data = {
        'id': diary_id,
    }

    data_response = GeneralDataResponse(
        code=status_code,
        status="OK",
        data=data,
    )

    return data_response  # FastAPI akan otomatis menangani serialisasi ke JSON
